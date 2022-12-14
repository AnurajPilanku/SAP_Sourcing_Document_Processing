#s=[{'Master_Agreement_ID': '2021-00009_MA', 'Folder_Name': '2021-00009_MA _ UV PURE TECHNOLOGIES INC _ 10001945', 'Master_Agreements': ['Test.docx', '1234567.docx'], 'Sub_Document Details': [{'2022-00039_AGR': ['1234567.docx', '10121314.xlsx', '20304050.txt']}, {'2022-00040_AGR': ['1234567.docx', '10121314.xlsx', '20304050.txt']}, {'2022-00047_AGR': ['1234567.docx', '10121314.xlsx', '20304050.txt']}]}, {'Master_Agreement_ID': '2021-00012_MA', 'Folder_Name': '2021-00012_MA _ SELECT ENGINEERING INC _ 1502857', 'Master_Agreements': 'EMPTY', 'Sub_Document Details': 'EMPTY'}, {'Master_Agreement_ID': '2021-00013_MA', 'Folder_Name': '2021-00013_MA _ BLACK BOX CANADA CORP _ 1501440', 'Master_Agreements': 'EMPTY', 'Sub_Document Details': 'EMPTY'}, {'Master_Agreement_ID': '2021-00013_MA', 'Folder_Name': '2021-00013_MA _ ARC ROYAL LIMITED _ 10000045', 'Master_Agreements': ['Test.docx'], 'Sub_Document Details': [{'2021-00004_AGR': 'Empty'}, {'2022-00042_AGR': 'Empty'}]}]
import openpyxl
import os
import pandas

#dwnloaddetailspath=r"C:\Users\2040664\anuraj\bvm\queryout\sd.xlsx"
#Maindownloaddetails=r"C:\Users\2040664\anuraj\bvm\queryout\main_sd.xlsx"
def Downloadfiles_Excel(s,downlodfilestorepath):
    dwnloaddetailspath=downlodfilestorepath.replace('SAP_summary.txt','temp.xlsx')
    Maindownloaddetails=downlodfilestorepath.replace('SAP_summary.txt','Empty_folder_Details.xlsx')
    Header = ['Master_Agreement_ID', 'Folder_Name', 'Master_Agreements', 'Sub_Document Details']
    try:
        if os.path.exists(Maindownloaddetails):
            main_wb=openpyxl.load_workbook(Maindownloaddetails)
            mainws = main_wb.active
        else:
            main_wb = openpyxl.Workbook()
            mainws=main_wb.active
            for hd in range(0,len(Header)):
                mainws.cell(column=hd+1,row=1).value=Header[hd]
        wb=openpyxl.Workbook()
        ws=wb.active
        for dic in s:
            for i in range(0,len(dic)):
                if list(dic.keys())[i] in ['Master_Agreement_ID','Folder_Name']:
                    ws.cell(row=s.index(dic)+1,column=i+1).value=dic[list(dic.keys())[i]]
                elif list(dic.keys())[i] in ['Master_Agreements']:
                    if type(dic[list(dic.keys())[i]])!=list:
                        if dic[list(dic.keys())[i]].lower()=='empty':
                            ws.cell(row=s.index(dic)+1,column=i+1).value = 'Empty'
                    else:
                        ws.cell(row=s.index(dic)+1, column=i+1).value = ','.join(dic[list(dic.keys())[i]])
                else:
                    if type(dic[list(dic.keys())[i]]) != list:
                        if dic[list(dic.keys())[i]].lower() == 'empty':
                            ws.cell(row=s.index(dic) + 1, column=i + 1).value = 'Empty'
                    else:
                        lsts=[ky for ky in dic[list(dic.keys())[i]]]
                        kys=[list(ky.keys())[0] for ky in lsts]
                        vls = [list(ky.values())[0] for ky in lsts]
                        if all(el=='Empty' for el in vls):
                            ws.cell(row=s.index(dic) + 1, column=i + 1).value = ','.join(kys)
                            if any(el == 'Empty' for el in vls):
                                ws.cell(row=s.index(dic) + 1, column=i + 1).value = ','.join(kys)
        wb.save(dwnloaddetailspath)
        wb.close()

        againwb=openpyxl.load_workbook(dwnloaddetailspath)
        agaiws=againwb.active
        againrowcount=agaiws.max_row
        mainrowcount=mainws.max_row
        for col in range(1,5):
            for row in range(1,againrowcount+1):
                if agaiws.cell(column=3, row=row).value =='Empty' or agaiws.cell(column=4, row=row).value not in [None,'',' ']:
                    mainws.cell(column=col,row=mainrowcount+row).value=agaiws.cell(column=col,row=row).value
        main_wb.save(Maindownloaddetails)
        main_wb.close()
        againwb.save(dwnloaddetailspath)
        againwb.close()
        os.remove(dwnloaddetailspath)
        return Maindownloaddetails
    except Exception as e:
        return e






