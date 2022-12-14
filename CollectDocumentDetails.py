import os
import openpyxl

def collect_Document_Details(folderName,path,singlelist):
    try:
        colhd=['MasterAgreementID','Effective_date','Expiration_date','SupplierName','ExternalID']
        folderpath=os.path.join(path,folderName)
        isExist=os.path.exists(folderpath)
        if isExist==True:
            wb=openpyxl.load_workbook(folderpath)
            ws=wb.active
            rc=ws.max_row
            for col in range(1,ws.max_column+1):
                ws.cell(column=col, row=rc+1).value = singlelist[col-1]
            wb.save(folderpath)
            wb.close()
        else:
            wb_=openpyxl.Workbook()
            ws_=wb_.active
            for ln in range(1,len(singlelist)+1):
                ws_.cell(row=1,column=ln).value=colhd[ln-1]
                ws_.cell(row=2,column=ln).value=singlelist[ln-1]
            wb_.save(folderpath)
            wb_.close()
        return "Data Addedd successfully for : "+singlelist[0]
    except Exception as e:
        return e
    
